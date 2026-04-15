"""PDF + XLSX exporters for decision reports.

Used by /api/export/decision-report.pdf and /api/export/decision-report.xlsx.
Both take a pre-filtered DataFrame + summary payload and return raw bytes.
"""
from datetime import datetime, timezone
from io import BytesIO

import pandas as pd


def _format_timestamp() -> str:
    return datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M UTC")


_PDF_COLUMNS = [
    "flight_id", "origin", "destination", "assigned_aircraft",
    "assigned_delay", "is_canceled", "decision_reason",
]

_XLSX_FLIGHT_COLUMNS = [
    "flight_id", "origin", "destination", "departure_time", "arrival_time",
    "assigned_aircraft", "crew_id", "assigned_delay", "is_canceled",
    "passenger_count", "decision_reason", "slot_pressure_flag",
]


def build_pdf_report(df: pd.DataFrame, summary: dict, filter_label: str) -> bytes:
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.platypus import (
        SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle,
    )

    buf = BytesIO()
    doc = SimpleDocTemplate(
        buf, pagesize=landscape(A4),
        leftMargin=24, rightMargin=24, topMargin=28, bottomMargin=24,
    )
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("Decision Report — Aviation Digital Twin", styles["Title"]))
    story.append(Paragraph(f"Generated: {_format_timestamp()}", styles["Normal"]))
    story.append(Paragraph(f"Filter: {filter_label}", styles["Normal"]))
    story.append(Spacer(1, 10))

    summary_rows = [["Metric", "Value"]]
    for key, value in summary.items():
        if isinstance(value, float):
            value = f"{value:,.2f}"
        summary_rows.append([str(key), str(value)])
    summary_table = Table(summary_rows, hAlign="LEFT")
    summary_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1f2937")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
        ("FONTSIZE", (0, 0), (-1, -1), 9),
    ]))
    story.append(summary_table)
    story.append(Spacer(1, 14))

    story.append(Paragraph("Top Flights (by cancellation / delay)", styles["Heading3"]))

    cols = [c for c in _PDF_COLUMNS if c in df.columns]
    table_data = [cols]
    preview = df.sort_values(
        by=["is_canceled", "assigned_delay"], ascending=[False, False]
    ).head(25) if not df.empty else df
    for _, row in preview.iterrows():
        table_data.append([str(row.get(c, ""))[:48] for c in cols])

    flights_table = Table(table_data, repeatRows=1)
    flights_table.setStyle(TableStyle([
        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#0f766e")),
        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
        ("GRID", (0, 0), (-1, -1), 0.2, colors.lightgrey),
        ("FONTSIZE", (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS", (0, 1), (-1, -1),
         [colors.whitesmoke, colors.HexColor("#f3f4f6")]),
    ]))
    story.append(flights_table)

    doc.build(story)
    return buf.getvalue()


def build_xlsx_report(df: pd.DataFrame, summary: dict, filter_label: str) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill

    wb = Workbook()

    ws_summary = wb.active
    ws_summary.title = "Summary"
    ws_summary.append(["Generated", _format_timestamp()])
    ws_summary.append(["Filter", filter_label])
    ws_summary.append([])
    ws_summary.append(["Metric", "Value"])
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="1F2937")
    for cell in ws_summary[4]:
        cell.font = header_font
        cell.fill = header_fill
    for key, value in summary.items():
        ws_summary.append([str(key), _xlsx_safe(value)])

    ws_flights = wb.create_sheet("Flights")
    flight_cols = [c for c in _XLSX_FLIGHT_COLUMNS if c in df.columns]
    ws_flights.append(flight_cols)
    for cell in ws_flights[1]:
        cell.font = header_font
        cell.fill = header_fill
    if not df.empty:
        export_df = df.sort_values(
            by=["is_canceled", "assigned_delay"], ascending=[False, False]
        )
        for _, row in export_df.iterrows():
            ws_flights.append([_xlsx_safe(row.get(c)) for c in flight_cols])

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _xlsx_safe(value):
    if value is None:
        return ""
    if isinstance(value, pd.Timestamp):
        return value.to_pydatetime().replace(tzinfo=None)
    if isinstance(value, (list, dict, tuple, set)):
        import json as _json
        try:
            return _json.dumps(value, ensure_ascii=False, default=str)
        except Exception:
            return str(value)
    if hasattr(value, "item"):
        try:
            return value.item()
        except Exception:
            return str(value)
    return value
